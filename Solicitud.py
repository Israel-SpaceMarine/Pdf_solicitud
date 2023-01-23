from io import BytesIO
from flask import Flask, request, send_file
import io
from os import remove
from flask_cors import CORS
import json
import os
import glob
from auth.jwt import check_token


from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)


app.config['TEMPLATES_AUTO_RELOAD'] = True


@app.route('/solicitud',  methods=['POST'])
@check_token

# funcion donde se recibe el json, se procesa y se remueven los xlsx para no tener basura en el cache
def get_data():
    data = request.data
    data = json.loads(data)
    for f in glob.iglob('*.xlsx', recursive=True):
        os.remove(f)
    return solicitud(data)

# funcion donde se abre la plantilla del excel donde se depositan los datos 
def solicitud(data):
    print(data)

    prueba = data["datos"]["Fecha"]
    print(prueba)

    Tipo_de_usuario = data["datos"]["AccesoApp"]
    

    # se lee la plantilla del excel para su llenado de informacion
    workbook = load_workbook(filename="./templates/template_sol_usuarios.xlsx")
    workbook.sheetnames
    sheet = workbook.active
    # llenado del excel
    x = data["datos"]["Fecha"]
    x = x.split("T")
    y =x[0]
    w = y.split("-")
    Fecha = w[2]+"/"+w[1]+"/"+w[0] 
    sheet["C9"] = Fecha
    sheet["C10"] = data["datos"]["TipoDeMovimiento"]
    sheet["C11"] = data["datos"]["Nombre"]
    sheet["C12"] = data["datos"]["ApellidoPaterno"]
    sheet["C13"] = data["datos"]["ApellidoMaterno"]
    sheet["C14"] = data["datos"]["NombreUsuario"]
    sheet["C15"] = data["datos"]["Correo"]
    sheet["C16"] = data["datos"]["CUPR"]
    sheet["C17"] = data["datos"]["RFC"]
    sheet["C18"] = data["datos"]["Telefono"]
    sheet["C19"] = data["datos"]["Extensión"]
    sheet["C20"] = data["datos"]["Celular"]
    sheet["C21"] = ""
    sheet["C22"] = data["datos"]["AccesoApp"]
    #sheet["A28"] = data["datos"]["Nota"] 
    # nombre del documento
    filename = Tipo_de_usuario + ".xlsx"
    file_stream = BytesIO()
    # se guarda el documetno
    workbook.save(Tipo_de_usuario + ".xlsx")
    

    file_stream.seek(0)
    
    # se envia el documento para su descarga
    return send_file("./" + filename , as_attachment=True, )
   



if __name__ == "__main__":

    app.run(host='0.0.0.0', port=90, debug=True)